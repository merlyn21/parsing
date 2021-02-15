#!/usr/bin/python3

import openpyxl

# создаем новый excel-файл
wb = openpyxl.Workbook()

# добавляем новый лист
wb.create_sheet(title = 'Первый лист', index = 0)

# получаем лист, с которым будем работать
sheet = wb['Первый лист']

for row in range(1, 4):
    for col in range(1, 4):
        value = str(row) + str(col)
        cell = sheet.cell(row = row, column = col)
        cell.value = value

wb.save('example.xlsx')
